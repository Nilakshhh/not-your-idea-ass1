from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import os

# Configuration for data extraction
data_config = [
    {
        "name": "Timing",
        "xpath": "//p//strong[contains(text(), 'Timing')] |  //p//b[contains(text(), 'Timing')]",
        "default": "Not Available"
    },
    {
        "name": "Address",
        "xpath": "//p//strong[contains(text(), 'Address')] | //p//b[contains(text(), 'Address')]",
        "default": "Not Available"
    },
    {
        "name": "Must Have",
        "xpath": "//p//strong[contains(text(),'Must have') or contains(text(),'Must-have') or contains(text(),'Must-Have') or contains(text(),'Must Have')] | //p//b[contains(text(),'Must have') or contains(text(),'Must-have') or contains(text(),'Must-Have') or contains(text(),'Must Have')]",
        "default": "Not Available"
    },
]

driver = webdriver.Chrome()
driver.maximize_window()

def get_url(url):
    driver.get(url)

def get_post_links():
    all_post_on_page_url = []
    all_post_on_page = driver.find_elements(By.XPATH, "//div[@class='elementor-widget-container']//article//a[@class='elementor-post__thumbnail__link']")
    for post in all_post_on_page:
        all_post_on_page_url.append(post.get_attribute("href"))
    return all_post_on_page_url

def get_all_pages_url():
    all_pages_url = []
    all_pages = driver.find_elements(By.XPATH, "//a[@class='page-numbers']")
    for pages in all_pages:
        all_pages_url.append(pages.get_attribute("href"))
    return all_pages_url

def get_data(field):
    all_data_list = []
    strong_list = driver.find_elements(By.XPATH, field["xpath"])

    if strong_list:
        for item in strong_list:
            parent = item.find_element(By.XPATH, "..")
            text = parent.text.replace(item.text, '').strip()
            all_data_list.append(text.replace(": ", '').replace('\u202f', ''))
        return all_data_list
    else:
        return [field["default"]]

def get_data_json():
    data = {"Link": driver.current_url}
    for field in data_config:
        data[field["name"]] = get_data(field)
    return data

def append_to_excel(data, filename='unsobered_data_scrapped.xlsx'):
    # Check if the file exists
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Write headers dynamically
        headers = ["Link"] + [field["name"] for field in data_config]
        ws.append(headers)

    # Get the maximum number of values for any field
    max_length = max(len(data[field["name"]]) for field in data_config)

    # Append rows for each entry
    for i in range(max_length):
        row = [data['Link']]  # Start with the Link
        for field in data_config:
            values = data[field["name"]]
            # Append the i-th value if it exists, otherwise append an empty string
            row.append(values[i] if i < len(values) else "")
        ws.append(row)

    # Save the workbook
    wb.save(filename)


def main():
    get_url("https://unsobered.com/category/in-your-city/")
    all_post_url = []
    all_pages_url = get_all_pages_url()
    all_post_url.extend(get_post_links())

    # Uncomment the loop below if you want to visit all pages
    for url in all_pages_url:
        get_url(url)
        all_post_url.extend(get_post_links())

    for url in all_post_url:
        get_url(url)
        data = get_data_json()
        append_to_excel(data)  # Append data to Excel

if __name__ == '__main__':
    main()